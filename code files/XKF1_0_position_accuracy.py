import numpy as np
import openpyxl
import matplotlib.pyplot as plt
import time

time_start = time.time()

wb = openpyxl.load_workbook("Single_point_data_testcase19_XKF1_0_python.xlsx")
sheet = wb.sheetnames
ws = wb[sheet[0]]

pos_ne = [0]*(ws.max_row - 1)
pos_ne_mean_temp = 0
pos_ne_sigma_temp = 0

del_t = 0.1    
n = 5

for row in range(1, ws.max_row):
    pos_ne_temp = 0
    for col in ws.iter_cols(0, ws.max_column):
        pos_ne_temp += round(col[row].value, n)**2
    pos_ne[row-1] = round(np.sqrt(pos_ne_temp), n)
    pos_ne_mean_temp += pos_ne[row-1]
pos_ne_mean = round(pos_ne_mean_temp / len(pos_ne), n)
    
for row in range(1, ws.max_row):
    pos_ne_sigma_temp += round((pos_ne[row-1] - pos_ne_mean)**2, n)
pos_ne_sigma = round(np.sqrt(pos_ne_sigma_temp / (len(pos_ne)-1)), n)

# Plotting commands
t = np.linspace(del_t,(ws.max_row-1)*del_t,ws.max_row-1)

params = {'figure.figsize': (8, 5), 
          'axes.labelsize': 16,
          'axes.titlesize': 16,
          'xtick.labelsize': 10, 
          'ytick.labelsize' : 10}
plt.rcParams.update(params)

plt.xlabel("Time (s)")
plt.ylabel("Position (m)")
plt.title("Position vs time")

plt.plot(t, pos_ne, color='b', label='pos_ne')

plt.legend(fontsize=12)
plt.xlim([0, max(t)])
plt.ylim([0, max(pos_ne)+0.5])

plt.show()

time_end = time.time()

print("Position mean =", pos_ne_mean, "m")
print("Position sigma =", pos_ne_sigma, "m")
print("Computation time =", round((time_end - time_start)/60, 3), "min")