import numpy as np
import openpyxl
import matplotlib.pyplot as plt
import time

time_start = time.time()

wb = openpyxl.load_workbook("Single_point_data_testcase21_IMU_2_python.xlsx")
sheet = wb.sheetnames
ws = wb[sheet[1]]

acc_x = [0]*(ws.max_row - 1)
acc_y = [0]*(ws.max_row - 1)
acc_xy = [0]*(ws.max_row - 1)
vel_x = [0]*(ws.max_row - 1)
vel_y = [0]*(ws.max_row - 1)
vel_xy = [0]*(ws.max_row - 1)
pos_x = [0]*(ws.max_row - 1)
pos_y = [0]*(ws.max_row - 1)
pos_xy = [0]*(ws.max_row - 1)

del_t = 0.04    
n = 5

for row in range(1, 2):
    for col in ws.iter_cols(0, 1):
        acc_x[row-1] = round(col[row].value, n)
    for col in ws.iter_cols(1, 2):
        acc_y[row-1] = round(col[row].value, n)
    acc_xy[row-1] = round(np.sqrt(acc_x[row-1]**2 + acc_y[row-1]**2), n)

for row in range(2, 3):
    for col in ws.iter_cols(0, 1):
        acc_x[row-1] = round(col[row].value, n)
        vel_x[row-2] = round(0.5*(acc_x[row-2] + acc_x[row-1])*del_t, n)
    for col in ws.iter_cols(1, 2):
        acc_y[row-1] = round(col[row].value, n)
        vel_y[row-2] = round(0.5*(acc_y[row-2] + acc_y[row-1])*del_t, n)
    acc_xy[row-1] = round(np.sqrt(acc_x[row-1]**2 + acc_y[row-1]**2), n)
    # vel_xy[row-2] = round(np.sqrt(vel_x[row-2]**2 + vel_y[row-2]**2), n)
    vel_xy[row-2] = round(0.5*(acc_xy[row-2] + acc_xy[row-1])*del_t, n)

for row in range(3, 4):
    for col in ws.iter_cols(0, 1):
        acc_x[row-1] = round(col[row].value, n)
        vel_x[row-2] = vel_x[row-3] + round(0.5*(acc_x[row-2] + acc_x[row-1])*del_t, n)
        pos_x[row-3] = round(0.5*(vel_x[row-3] + vel_x[row-2])*del_t, n)
    for col in ws.iter_cols(1, 2):
        acc_y[row-1] = round(col[row].value, n)
        vel_y[row-2] = vel_y[row-3] + round(0.5*(acc_y[row-2] + acc_y[row-1])*del_t, n)
        pos_y[row-3] = round(0.5*(vel_y[row-3] + vel_y[row-2])*del_t, n)
    acc_xy[row-1] = round(np.sqrt(acc_x[row-1]**2 + acc_y[row-1]**2), n)
    # vel_xy[row-2] = round(np.sqrt(vel_x[row-2]**2 + vel_y[row-2]**2), n)
    # pos_xy[row-3] = round(np.sqrt(pos_x[row-3]**2 + pos_y[row-3]**2), n)
    vel_xy[row-2] = vel_xy[row-3] + round(0.5*(acc_xy[row-2] + acc_xy[row-1])*del_t, n)
    pos_xy[row-3] = round(0.5*(vel_xy[row-3] + vel_xy[row-2])*del_t, n)

for row in range(4, ws.max_row):
    for col in ws.iter_cols(0, 1):
        acc_x[row-1] = round(col[row].value, n)
        vel_x[row-2] = vel_x[row-3] + round(0.5*(acc_x[row-2] + acc_x[row-1])*del_t, n)
        pos_x[row-3] = pos_x[row-4] + round(0.5*(vel_x[row-3] + vel_x[row-2])*del_t, n)
    for col in ws.iter_cols(1, 2):
        acc_y[row-1] = round(col[row].value, n)
        vel_y[row-2] = vel_y[row-3] + round(0.5*(acc_y[row-2] + acc_y[row-1])*del_t, n)
        pos_y[row-3] = pos_y[row-4] + round(0.5*(vel_y[row-3] + vel_y[row-2])*del_t, n)
    acc_xy[row-1] = round(np.sqrt(acc_x[row-1]**2 + acc_y[row-1]**2), n)
    # vel_xy[row-2] = round(np.sqrt(vel_x[row-2]**2 + vel_y[row-2]**2), n)
    # pos_xy[row-3] = round(np.sqrt(pos_x[row-3]**2 + pos_y[row-3]**2), n)
    vel_xy[row-2] = vel_xy[row-3] + round(0.5*(acc_xy[row-2] + acc_xy[row-1])*del_t, n)
    pos_xy[row-3] = pos_xy[row-4] + round(0.5*(vel_xy[row-3] + vel_xy[row-2])*del_t, n)

# Plotting commands
t = np.linspace(del_t,(ws.max_row-1)*del_t,ws.max_row-1)

params = {'figure.figsize': (8, 5), 
          'axes.labelsize': 16,
          'axes.titlesize': 16,
          'xtick.labelsize': 10, 
          'ytick.labelsize' : 10}
plt.rcParams.update(params)

plt.xlabel("Time (s)")
plt.ylabel("Accelration (m/s/s)")
plt.title("Acceleration vs time")

plt.plot(t, acc_xy, color='b', label='acc_xy')

plt.legend(fontsize=12)
# plt.xlim([0, max(t)])
# plt.ylim([0, max(acc_xy)])

plt.show()

plt.xlabel("Time (s)")
plt.ylabel("Velocity (m/s)")
plt.title("Velocity vs time")

plt.plot(t, vel_xy, color='b', label='vel_xy')

plt.legend(fontsize=12)
# plt.xlim([0, max(t)])
# plt.ylim([0, max(vel_xy)])

plt.show()

plt.xlabel("Time (s)")
plt.ylabel("Position (m)")
plt.title("Position vs time")

plt.plot(t, pos_xy, color='b', label='pos_xy')

plt.legend(fontsize=12)
# plt.xlim([0, max(t)])
# plt.ylim([0, max(pos_xy)])

plt.show()

time_end = time.time()

print("Computation time =", round((time_end - time_start)/60, 3), "min")