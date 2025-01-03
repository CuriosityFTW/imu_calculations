import numpy as np
import openpyxl
import matplotlib.pyplot as plt
import time

time_start = time.time()

wb = openpyxl.load_workbook("Single_point_data_testcase21_IMU_2_python.xlsx")
sheet = wb.sheetnames
ws = wb[sheet[1]]

acc = [0]*(ws.max_row - 1)
acc_mean_temp = 0
acc_sigma_temp = 0

del_t = 0.04    
n = 3

for row in range(1, ws.max_row):
    for col in ws.iter_cols(2, ws.max_column):
        acc[row-1] = round(col[row].value, n)
    acc_mean_temp += acc[row-1]
acc_mean = round(acc_mean_temp / len(acc), n)
    
for row in range(1, ws.max_row):
    acc_sigma_temp += round((acc[row-1] - acc_mean)**2, n)
acc_sigma = round(np.sqrt(acc_sigma_temp / (len(acc)-1)), n)
acc_noise = round(acc_sigma**2, 10)

# Plotting commands
t = np.linspace(del_t,(ws.max_row-1)*del_t,ws.max_row-1)

params = {'figure.figsize': (8, 5), 
          'axes.labelsize': 16,
          'axes.titlesize': 16,
          'xtick.labelsize': 10, 
          'ytick.labelsize' : 10}
plt.rcParams.update(params)

plt.xlabel("Time (s)")
plt.ylabel("Acceleration (m/s/s)")
plt.title("Acceleration vs time")

plt.plot(t, acc, color='b', label='acc')

plt.legend(fontsize=12)
plt.xlim([0, max(t)])
plt.ylim([min(acc)-0.1, max(acc)+0.1])

plt.show()

time_end = time.time()

print("Acceleration mean =", acc_mean, "m/s/s")
print("Accelerometer RMS noise =", acc_sigma, "m/s/s")
print("Accelerometer process noise =", acc_noise, "m2/s4")
print("Computation time =", round((time_end - time_start)/60, 3), "min")